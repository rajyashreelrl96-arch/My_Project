import openpyxl
from datetime import datetime
from selenium.webdriver.common.by import By
from orange_hrm import LoginPage

def test_data_driven(setup):
    driver = setup
    lp = LoginPage(driver)
    file = "D:/Rajalakshi-GUVI learnings/Orangehrm.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):  # skip header row
        username = sheet.cell(row, 2).value
        password = sheet.cell(row, 3).value

        lp.enter_username(username)
        lp.enter_password(password)
        lp.click_login()

        # Write date and time
        date = datetime.now().strftime("%d/%m/%y")
        time = datetime.now().strftime("%H:%M")
        sheet.cell(row, 4).value = date
        sheet.cell(row, 5).value = time

        # Check login result
        if lp.is_login_success():
            sheet.cell(row, 7).value = "Passed"
            driver.find_element(By.XPATH, "//span[@class='oxd-userdropdown-tab']/i").click()
            driver.find_element(By.XPATH, "//a[.='Logout']").click()
            lp = LoginPage(driver)  # reset page object
        else:
            sheet.cell(row, 7).value = "Failed"
            driver.refresh()

    # Save workbook back to file
    workbook.save(file)